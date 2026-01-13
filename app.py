import os
import streamlit as st
import yaml
import openpyxl
import collections
from io import BytesIO, StringIO
import re
import pandas as pd
from ruamel.yaml import YAML
from ruamel.yaml.comments import CommentedMap

# -----------------------------
# Helpers for resource normalization
# -----------------------------
def canon(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def normalize_cpu(val):
    if not val:
        return None
    s = str(val).lower().strip()
    if s.endswith("m"):
        try:
            return float(s[:-1]) / 1000
        except:
            return None
    try:
        return float(s)
    except:
        return None

def normalize_mem(val):
    if not val:
        return None
    s = str(val).lower().strip()
    try:
        if s.endswith("gi") or s.endswith("g"):
            return float(s[:-2] if s.endswith("gi") else s[:-1])
        if s.endswith("mi") or s.endswith("m"):
            return float(s[:-2] if s.endswith("mi") else s[:-1]) / 1024
        if s.endswith("ki") or s.endswith("k"):
            return float(s[:-2] if s.endswith("ki") else s[:-1]) / (1024 * 1024)
        return float(s) / (1024 * 1024 * 1024)
    except:
        return None

def get_case_insensitive(d, key):
    if not isinstance(d, dict):
        return None
    for k, v in d.items():
        if str(k).lower() == key.lower():
            return v
    return None

def extract_release_from_filename(filename: str):
    match = re.search(r"(\d+\.\d+\.\d+)", filename)
    if match:
        return match.group(1)
    return None

# -----------------------------
# Flatten YAML (anchors included)
# -----------------------------
yaml_loader = YAML()
yaml_loader.preserve_quotes = True
yaml_loader.allow_duplicate_keys = True

def flatten_yaml_with_inline_anchors(data, parent_key='', sep='.'):
    items = []
    if isinstance(data, CommentedMap):
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            items.extend(flatten_yaml_with_inline_anchors(v, new_key, sep=sep))
    elif isinstance(data, list):
        for idx, item in enumerate(data):
            new_key = f"{parent_key}{sep}{idx}"
            items.extend(flatten_yaml_with_inline_anchors(item, new_key, sep=sep))
    else:
        anchor = getattr(data, 'anchor', None)
        anchor_value = getattr(anchor, 'value', None)
        if data is None and anchor_value:
            value_str = f"&{anchor_value}"
        else:
            if isinstance(data, bool):
                val = "true" if data else "false"
            else:
                val = data if isinstance(data, str) else str(data)
            value_str = f"&{anchor_value} {val}" if anchor_value else val
        items.append((parent_key, value_str))
    return items

def process_yaml_to_df(yaml_data):
    flat_records = []
    for main_key, value in yaml_data.items():
        flattened = flatten_yaml_with_inline_anchors(value)
        for param, val in flattened:
            flat_records.append({
                'Main Name': main_key,
                'Parameter': param,
                'Value': val
            })
    return pd.DataFrame(flat_records)

# -----------------------------
# Helpers for fuzzy matching in Excel
# -----------------------------
def find_matching_sheet(wb, target_name):
    # Flexible matching: ignore case and common separators
    target_norm = target_name.strip().lower().replace("_", "").replace("-", "").replace(" ", "")
    for sheet in wb.sheetnames:
        sheet_norm = sheet.strip().lower().replace("_", "").replace("-", "").replace(" ", "")
        if sheet_norm == target_norm:
            return wb[sheet]
    return None

def find_section_start(ws, target_name):
    target_norm = target_name.strip().lower()
    for r in range(1, ws.max_row + 1):
        row_text = " ".join([str(c.value).lower() for c in ws[r] if c.value])
        if target_norm in row_text:
            return r
    return None


# -----------------------------
# Core Excel processor
# -----------------------------
def process_yaml_to_excel(yaml_dict, release_map, nf_type, flat_dfs, template_file_buffer):
    wb = openpyxl.load_workbook(template_file_buffer)

    section_map = {
        "SCP": {"scp": "SCP PODS", "cndbtier": "SCP cnDBTier PODS"},
        "NRF": {"nrf": "NRF PODS", "cndbtier": "NRF cnDBTier PODS"},
        "PCF": {
            "data": "PCF DATA PODS",
            "voice": "PCF VOICE PODS",
            "cndbtier": "PCF cnDBTier PODS" # Added mapping for pods
        }
    }

    current_section_map = section_map.get(nf_type, {})
    yaml_dict = {canon(k): v for k, v in yaml_dict.items()}
    preview_data = {}

    def fill_section(section_name, yaml_data, release_value):
        ws = None
        start_row = None
        for sheet in wb.worksheets:
            row_num = find_section_start(sheet, section_name)
            if row_num:
                ws = sheet
                start_row = row_num
                break
        if not ws or not start_row:
            return pd.DataFrame()

        header_row = None
        for r in range(start_row, ws.max_row + 1):
            row_vals = [str(c.value).strip().lower() if c.value else "" for c in ws[r]]
            if "name" in row_vals:
                header_row = r
                break
        if not header_row:
            return pd.DataFrame()

        headers = [c.value for c in ws[header_row]]
        h_to_idxs = collections.defaultdict(list)
        for idx, h in enumerate(headers):
            if h:
                h_to_idxs[h.strip().lower()].append(idx)

        def first_header_idx(label: str):
            idxs = h_to_idxs.get(label.strip().lower(), [])
            return idxs[0] if idxs else None

        name_idx = first_header_idx("name")
        cpu_req_idx = first_header_idx("cpu request per pod (#)")
        cpu_lim_idx = first_header_idx("cpu limit per pod (#)")
        mem_req_idx = first_header_idx("memory request per pod (gb)")
        mem_lim_idx = first_header_idx("memory limit per pod (gb)")
        min_rep_idx = first_header_idx("min # replicas")
        max_rep_idx = first_header_idx("max # replicas")
        pvc_count_idx = first_header_idx("count of pvcs")
        storage_idx = first_header_idx("storage pvc data disk (gib)")
        release_idx = first_header_idx("release introduced")

        global_replicas = yaml_data.get("global", {}) if "global" in yaml_data else {}

        yaml_keys = [k for k in yaml_data.keys() if isinstance(yaml_data[k], dict) and k != "global"]
        yaml_iter = iter(yaml_keys)
        collected = []
        release_written = False

        for row in ws.iter_rows(min_row=header_row + 1, max_col=len(headers)):
            name_cell = row[name_idx] if name_idx is not None else None
            if not name_cell:
                continue
            if str(name_cell.value).strip().lower().startswith("total"):
                break

            if not name_cell.value:
                try:
                    comp_key = next(yaml_iter)
                    name_cell.value = comp_key
                except StopIteration:
                    continue
            else:
                comp_key = next((k for k in yaml_keys if canon(k) == canon(name_cell.value)), None)

            if not comp_key:
                continue

            comp = yaml_data.get(comp_key, {})
            resources = comp.get("resources", {}) or {}
            req = resources.get("requests", {}) or {}
            lim = resources.get("limits", {}) or {}

            if cpu_req_idx is not None: row[cpu_req_idx].value = normalize_cpu(req.get("cpu"))
            if cpu_lim_idx is not None: row[cpu_lim_idx].value = normalize_cpu(lim.get("cpu"))
            if mem_req_idx is not None: row[mem_req_idx].value = normalize_mem(req.get("memory"))
            if mem_lim_idx is not None: row[mem_lim_idx].value = normalize_mem(lim.get("memory"))

            min_reps = get_case_insensitive(comp, "minReplicas")
            max_reps = get_case_insensitive(comp, "maxReplicas")

            if section_name.lower().find("cndbtier") != -1 and global_replicas:
                for g_key, g_val in global_replicas.items():
                    if not isinstance(g_val, (int, float)):
                        continue
                    key_lower = g_key.lower()
                    if "replica" not in key_lower:
                        continue
                    comp_name = key_lower.replace("replicacount", "").replace("replicamaxcount", "")
                    if canon(comp_key) == canon(comp_name):
                        if key_lower.endswith("maxcount"):
                            max_reps = g_val
                        else:
                            min_reps = g_val

            if min_rep_idx is not None: row[min_rep_idx].value = min_reps
            if max_rep_idx is not None: row[max_rep_idx].value = max_reps

            pvc_data = get_case_insensitive(comp, "pvc") or {}
            pvc_count_val = comp.get("pvcCount") or comp.get("pvc_count") or pvc_data.get("count")
            storage_val = (resources.get("storage") or comp.get("storage") or pvc_data.get("disksize") or pvc_data.get("size"))
            if pvc_count_idx is not None: row[pvc_count_idx].value = pvc_count_val
            if storage_idx is not None: row[storage_idx].value = normalize_mem(storage_val)

            if release_value and release_idx is not None and not release_written:
                row[release_idx].value = release_value
                release_written = True

            collected.append({
                "Excel Name": name_cell.value,
                "YAML Key": comp_key,
                "CPU Req": normalize_cpu(req.get("cpu")),
                "CPU Lim": normalize_cpu(lim.get("cpu")),
                "Mem Req (GB)": normalize_mem(req.get("memory")),
                "Mem Lim (GB)": normalize_mem(lim.get("memory")),
                "Min Replicas": min_reps,
                "Max Replicas": max_reps,
            })

        return pd.DataFrame(collected)

    for key, section in current_section_map.items():
        if key in yaml_dict:
            release_val = release_map.get(key)
            section_df = fill_section(section, yaml_dict[key], release_val)
            if not section_df.empty:
                preview_data[key] = section_df

    # -----------------------------
    # Fill Custom YAML sheets (REFINED MAPPING)
    # -----------------------------
    custom_sheet_map = {
        "SCP": {"scp": "SCP_Custom_yaml", "cndbtier": "SCP_cnDBTier_custom_yaml"},
        "NRF": {"nrf": "NRF_Custom_yaml", "cndbtier": "NRF_cnDBTier_custom_yaml"},
        "PCF": {
            "data": "PCF_Custom_yaml-Data",
            "voice": "PCF_Custom_yaml-Voice",
            "cndbtier": "PCF_cnDBTier_custom_yaml" # Added for PCF
        }
    }

    for key, sheet_name in custom_sheet_map.get(nf_type, {}).items():
        if key in flat_dfs:
            ws_custom = find_matching_sheet(wb, sheet_name)
            if ws_custom:
                for i, row in flat_dfs[key].iterrows():
                    ws_custom.cell(row=i + 2, column=1, value=row['Main Name'])
                    ws_custom.cell(row=i + 2, column=2, value=row['Parameter'])
                    ws_custom.cell(row=i + 2, column=3, value=row['Value'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, preview_data

# -----------------------------
# Streamlit UI (Restored to original format)
# -----------------------------
st.title("NAPD Automation Tool")
st.markdown("---")
st.header("1. Select Network Function Type")

nf_type = st.selectbox(
    "Choose the NF type you want to process:",
    ("SCP", "NRF", "PCF"),
    index=0
)

st.markdown("---")
st.header(f"2. Upload {nf_type} YAML Files")

yaml_files = {}
if nf_type == "SCP":
    scp_file = st.file_uploader("Upload SCP YAML", type=["yaml", "yml"])
    cndbtier_file = st.file_uploader("Upload cnDBTier YAML", type=["yaml", "yml"])
    if scp_file: yaml_files["scp"] = scp_file
    if cndbtier_file: yaml_files["cndbtier"] = cndbtier_file
elif nf_type == "NRF":
    nrf_file = st.file_uploader("Upload NRF YAML", type=["yaml", "yml"])
    cndbtier_file = st.file_uploader("Upload cnDBTier YAML", type=["yaml", "yml"])
    if nrf_file: yaml_files["nrf"] = nrf_file
    if cndbtier_file: yaml_files["cndbtier"] = cndbtier_file
elif nf_type == "PCF":
    pcf_data_file = st.file_uploader("Upload PCF DATA YAML", type=["yaml","yml"])
    pcf_voice_file = st.file_uploader("Upload PCF VOICE YAML", type=["yaml","yml"])
    pcf_cndbtier_file = st.file_uploader("Upload PCF cnDBTier YAML", type=["yaml","yml"]) # Added
    if pcf_data_file: yaml_files["data"] = pcf_data_file
    if pcf_voice_file: yaml_files["voice"] = pcf_voice_file
    if pcf_cndbtier_file: yaml_files["cndbtier"] = pcf_cndbtier_file # Key mapped to "cndbtier"


st.markdown("---")
st.header("3. Generate Excel File")

if st.button("Generate Excel"):
    if not yaml_files:
        st.error(f"Please upload at least one YAML file for {nf_type}.")
    else:
        yaml_dict = {}
        release_map = {}
        flat_dfs = {}

        if nf_type == "SCP":
            template_path = "scp_template.xlsx"
        elif nf_type == "NRF":
            template_path = "nrf_template.xlsx"
        elif nf_type == "PCF":
            template_path = "pcf_template.xlsx"

        try:
            with open(template_path, "rb") as f:
                template_buffer = f.read()
        except FileNotFoundError:
            st.error(f"Template file '{template_path}' not found in backend.")
            template_buffer = None

        if template_buffer:
            for name, f in yaml_files.items():
                try:
                    raw_text = f.read().decode("utf-8")
                    yaml_data = yaml_loader.load(StringIO(raw_text))
                    yaml_dict[name] = yaml_data
                    release_map[name] = extract_release_from_filename(f.name)
                    flat_dfs[name] = process_yaml_to_df(yaml_data)
                except Exception as e:
                    st.error(f"Error processing {f.name}: {e}")

            if yaml_dict:
                output, previews = process_yaml_to_excel(yaml_dict, release_map, nf_type, flat_dfs, BytesIO(template_buffer))
                
                main_nf_key = "data" if nf_type == "PCF" else nf_type.lower()
                release_number = release_map.get(main_nf_key, "Unknown")
                file_name = f"NAPD_{nf_type}_{release_number}.xlsx"

                st.markdown("---")
                st.header("4. Preview and Download")
                for key, df in previews.items():
                    if not df.empty:
                        st.subheader(f"Preview → {key.upper()} Section")
                        st.dataframe(df)

                st.download_button(
                    label="⬇️ Download Excel",
                    data=output,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8501))
